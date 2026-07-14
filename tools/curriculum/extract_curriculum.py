#!/usr/bin/env python3
"""Faithful extraction of the الراسخون Quran curriculum from the source spreadsheets.

Design contract (see tools/curriculum/README section in the report):

  * STRUCTURE decides what a row is. The Arabic text is a LABEL that we store
    verbatim and use only to CROSS-CHECK the structural decision.
  * Nothing is ever inferred from a session number, and nothing is ever
    fabricated. Every emitted session carries a `source` provenance record
    pointing at the exact file / sheet / row it came from.
  * The script VALIDATES first and writes nothing unless validation passes.
    `--write` emits JSON only after a clean report.

Usage:
    python extract_curriculum.py            # validate only, print report
    python extract_curriculum.py --write    # validate, then emit data/curriculum/*.json
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]          # al_rasikhoon/
OUTPUT_DIR = REPO_ROOT / "data" / "curriculum"


class Abort(Exception):
    pass


def _find_curriculum_dir() -> Path:
    """The read-only source tree. Never written to.

    Normally `../curriculum` next to the repo. When the extractor is run from a
    git worktree (`al_rasikhoon/.claude/worktrees/<x>/`) that sibling does not
    exist, so we walk up until we find the real one. `AL_RASIKHOON_CURRICULUM_DIR`
    overrides both.
    """
    override = os.environ.get("AL_RASIKHOON_CURRICULUM_DIR")
    if override:
        return Path(override).resolve()
    for base in [REPO_ROOT, *REPO_ROOT.parents]:
        candidate = base.parent / "curriculum"
        if candidate.is_dir():
            return candidate
    raise Abort("cannot locate the read-only curriculum/ source tree")


CURRICULUM_DIR = _find_curriculum_dir()

# --------------------------------------------------------------------------
# Level definitions
# --------------------------------------------------------------------------
# A level owns 3 juz. Juz are taught in DESCENDING order in levels 1-9
# (level 1 teaches 30, then 29, then 28).
#
# LEVEL 10 IS ASCENDING IN THE SOURCE (juz 1, then 2, then 3). This is not a
# guess: juz 1 carries no cumulative pair, juz 2's cumulative covers
# "البقرة 1 : 252" (= juz 1 + 2) and juz 3's covers "البقرة 1 : 286"
# (= juz 1 + 2 + 3). Cumulative scope strictly grows in that order, so that is
# the teaching order.
#
# THIS IS NOT MACHINE-VERIFIED FOR EVERY LEVEL. build_juz_sessions() only
# cross-checks a juz-/cumulative-tier label against the declared teaching
# order when the label actually NAMES a juz (JUZ_WORD_RE: "الجزء" /
# "الجزئين" / "الأجزاء") — true for levels 1-2 only, whose unit is a hizb.
# Levels 3-10 label their juz- and cumulative-tier assessments by SURAH
# ("من أول ... إلى أخر ...") or, for level 10, by verse number — never by juz
# number — so that cross-check is silently skipped for them (see
# test_level_3_cumulative_labels_name_no_juz). The derivation above for level
# 10 (and the descending order asserted for 1-9) was verified BY HAND against
# the source and is recorded here for whoever next touches this file; no code
# in this module re-derives or re-checks it.
LEVEL_NAMES_AR = {
    1: "المستوى الأول",
    2: "المستوى الثاني",
    3: "المستوى الثالث",
    4: "المستوى الرابع",
    5: "المستوى الخامس",
    6: "المستوى السادس",
    7: "المستوى السابع",
    8: "المستوى الثامن",
    9: "المستوى التاسع",
    10: "المستوى العاشر",
}

LEVEL_NAMES_EN = {
    1: "Level 1", 2: "Level 2", 3: "Level 3", 4: "Level 4", 5: "Level 5",
    6: "Level 6", 7: "Level 7", 8: "Level 8", 9: "Level 9", 10: "Level 10",
}

LEVEL_FOLDERS = {
    "المستوى الأول جاهز للمنسق": 1,
    "المستوى الثاني جاهز للمنسق": 2,
    "المستوى الثالث جاهز للمنسق": 3,
    "المستوى الرابع جاهز للمنسق": 4,
    "المستوى الخامس جاهز للمنسق": 5,
    "المستوى السادس جاهز للمنسق": 6,
    "المستوى السابع جاهز للمنسق": 7,
    "المستوى الثامن جاهز للمنسق": 8,
    "المستوى التاسع جاهز للمنسق": 9,
    "المستوى العاشر جاهز للمنسق": 10,
}


def juz_teaching_order(level: int) -> list[int]:
    """Juz numbers of a level, in TEACHING order."""
    if level == 10:
        return [1, 2, 3]  # source-verified ascending; see note above
    return [33 - 3 * level, 32 - 3 * level, 31 - 3 * level]


# --------------------------------------------------------------------------
# Documented source defects / deviations.
# An UNLISTED violation of a structural invariant aborts the run.
# --------------------------------------------------------------------------
@dataclass(frozen=True)
class Exception_:
    reason: str


# keyed by (level, juz)
EXPECTED_EXCEPTIONS: dict[tuple[int, int], dict[str, Any]] = {
    (10, 3): {
        "unit_pairs": 1,
        "juz_pairs": 0,
        "cumulative_pairs": 1,
        "second_pair_tier": "cumulative",
        "reason": (
            "The colour-filled tab of juz 3 of level 10 has a single teaching block "
            "(سورة البقرة 253:286, 4 new-memorisation lessons then 2 review-only ones) "
            "followed by ONE unit pair over البقرة 253:286 and then a pair over the whole "
            "of سورة البقرة 1:286 (= juz 1+2+3), i.e. the level cumulative. There is no "
            "second unit block and no juz-tier pair."
        ),
    },
}

# Raw (as-printed) session numbering that is not a clean +1 run. Renumbering is
# dense regardless; these are the ones we tolerate instead of aborting.
#
# Empty since the extractor started reading the COLOUR-FILLED tabs: the two gaps
# that used to live here (level 10 juz 2, 7 -> 11 and 16 -> 19) were defects of
# the abandoned draft tabs. Every authoritative tab numbers its sessions in a
# clean +1 run. The mechanism stays; there is simply nothing to excuse.
EXPECTED_NUMBERING_GAPS: dict[tuple[int, int], str] = {}

# Level 2's in-sheet hizb markers contradict the workbook they sit in (the file
# whose content is hizb 53 carries 'سرد الحزب رقم 54' and vice versa). We take the
# STRUCTURE (filename + juz-pair-comes-last) as truth and report the contradiction.
KNOWN_HIZB_LABEL_DEFECT_LEVELS = {2}

# --------------------------------------------------------------------------
# Cell / text helpers
# --------------------------------------------------------------------------
PLACEHOLDER_RE = re.compile(r"^[ـ_\-\s\.]*$")  # tatweel / underscore / dash runs
MARKER_RE = re.compile(r"^\s*(سرد|اختبار)")
HIZB_IN_TEXT_RE = re.compile(r"الحزب\s*(?:رقم\s*)?(\d+)")
JUZ_IN_TEXT_RE = re.compile(r"(\d+)")
LEVEL_WORD_RE = re.compile(r"المستوى")
JUZ_WORD_RE = re.compile(r"الجزء|الجزئين|الأجزاء")


def norm_cell(value: Any) -> Optional[str]:
    """Normalise a raw cell to a string, or None for empty/placeholder cells."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    if PLACEHOLDER_RE.match(text):
        return None
    return text


def as_int(value: Any) -> Optional[int]:
    text = norm_cell(value)
    if text is None:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def is_marker(text: Optional[str]) -> bool:
    return bool(text) and bool(MARKER_RE.match(text)) and len(text) > 6


# --------------------------------------------------------------------------
# The 114 surahs, in mushaf order. Every surah name the source names must be one
# of these — a name that is not is a TYPO IN THE SOURCE. We report it and store
# it verbatim; we never silently "correct" curriculum content.
# --------------------------------------------------------------------------
SURAH_NAMES: tuple[str, ...] = (
    "الفاتحة", "البقرة", "آل عمران", "النساء", "المائدة", "الأنعام", "الأعراف", "الأنفال",
    "التوبة", "يونس", "هود", "يوسف", "الرعد", "إبراهيم", "الحجر", "النحل", "الإسراء",
    "الكهف", "مريم", "طه", "الأنبياء", "الحج", "المؤمنون", "النور", "الفرقان", "الشعراء",
    "النمل", "القصص", "العنكبوت", "الروم", "لقمان", "السجدة", "الأحزاب", "سبأ", "فاطر",
    "يس", "الصافات", "ص", "الزمر", "غافر", "فصلت", "الشورى", "الزخرف", "الدخان", "الجاثية",
    "الأحقاف", "محمد", "الفتح", "الحجرات", "ق", "الذاريات", "الطور", "النجم", "القمر",
    "الرحمن", "الواقعة", "الحديد", "المجادلة", "الحشر", "الممتحنة", "الصف", "الجمعة",
    "المنافقون", "التغابن", "الطلاق", "التحريم", "الملك", "القلم", "الحاقة", "المعارج",
    "نوح", "الجن", "المزمل", "المدثر", "القيامة", "الإنسان", "المرسلات", "النبأ",
    "النازعات", "عبس", "التكوير", "الإنفطار", "المطففين", "الإنشقاق", "البروج", "الطارق",
    "الأعلى", "الغاشية", "الفجر", "البلد", "الشمس", "الليل", "الضحى", "الشرح", "التين",
    "العلق", "القدر", "البينة", "الزلزلة", "العاديات", "القارعة", "التكاثر", "العصر",
    "الهمزة", "الفيل", "قريش", "الماعون", "الكوثر", "الكافرون", "النصر", "المسد",
    "الإخلاص", "الفلق", "الناس",
)

_DIACRITICS_RE = re.compile(r"[ً-ْـ]")  # harakat, sukun, tatweel


def normalise_surah(name: str) -> str:
    """Fold the orthographic variation the source legitimately uses.

    Hamza carriers and ta-marbuta are written inconsistently across the
    workbooks (سبأ/سبا, الانشقاق/الإنشقاق) and both spellings are the same surah.
    WHITESPACE IS NOT COLLAPSED AWAY: 'المعار ج' is a typo, not a variant, and
    must not be normalised into 'المعارج'.
    """
    text = _DIACRITICS_RE.sub("", name)
    for src, dst in (("أ", "ا"), ("إ", "ا"), ("آ", "ا"), ("ٱ", "ا"),
                     ("ة", "ه"), ("ى", "ي"), ("ئ", "ي"), ("ؤ", "و")):
        text = text.replace(src, dst)
    return " ".join(text.split())


KNOWN_SURAH_NAMES = {normalise_surah(n) for n in SURAH_NAMES}


def is_known_surah(name: str) -> bool:
    return normalise_surah(name) in KNOWN_SURAH_NAMES


# --------------------------------------------------------------------------
# Sheet parsing
# --------------------------------------------------------------------------
@dataclass
class RawRow:
    file: str
    sheet: str
    row: int  # 0-based sheet row index
    raw_session_number: Optional[int]
    marker_text: Optional[str]
    marker_col: Optional[int]
    current: Optional[dict]
    recent: Optional[dict]
    distant: Optional[dict]


BLOCK_HEADERS = {
    "current": "مقرر المستوى الحالي",
    "recent": "مقرر تثبيت المستوى الحالي",
    "distant": "مقرر تثبيت المستوىات السابقة",
}
SESSION_HEADER = "الحصة رقم"


def locate_columns(df: pd.DataFrame) -> Optional[dict]:
    """Find the header row and the column index of each logical block."""
    for idx in range(min(12, len(df))):
        row = df.iloc[idx]
        cells = {j: norm_cell(v) for j, v in enumerate(row)}
        session_col = next((j for j, t in cells.items() if t and SESSION_HEADER in t), None)
        if session_col is None:
            continue
        cols = {"header_row": idx, "session": session_col}
        for key, label in BLOCK_HEADERS.items():
            col = next(
                (j for j, t in cells.items() if t and t.replace(" ", "").startswith(label.replace(" ", ""))),
                None,
            )
            if col is None:
                # tolerate the 'المستوىات' typo variants
                col = next((j for j, t in cells.items() if t and label[:12] in t), None)
            if col is None:
                return None
            cols[key] = col
        return cols
    return None


def read_block(row: pd.Series, start: int, exclude_col: Optional[int]) -> Optional[dict]:
    """A content block is 4 consecutive cells: from_surah, from_verse, to_surah, to_verse."""
    def cell(offset: int):
        col = start + offset
        if col >= len(row) or col == exclude_col:
            return None
        return row.iloc[col]

    from_surah = norm_cell(cell(0))
    from_verse = as_int(cell(1))
    to_surah = norm_cell(cell(2))
    to_verse = as_int(cell(3))
    if from_surah is None or to_surah is None or from_verse is None or to_verse is None:
        return None
    if is_marker(from_surah) or is_marker(to_surah):
        return None
    return {
        "from_surah": from_surah,
        "from_verse": from_verse,
        "to_surah": to_surah,
        "to_verse": to_verse,
    }


def parse_sheet(path: Path, sheet: str, df: pd.DataFrame, anomalies: list[str]) -> list[RawRow]:
    cols = locate_columns(df)
    if cols is None:
        return []

    title = " ".join(
        t for t in (norm_cell(v) for v in df.iloc[: cols["header_row"]].to_numpy().ravel()) if t
    )
    if "لاغي" in title:
        anomalies.append(
            f"{path.name} / {sheet}: sheet is stamped 'لاغي' (void) in its title — skipped"
        )
        return []

    rows: list[RawRow] = []
    for idx in range(cols["header_row"] + 1, len(df)):
        row = df.iloc[idx]
        session_num = as_int(row.iloc[cols["session"]]) if cols["session"] < len(row) else None

        marker_col = None
        marker_text = None
        for j, value in enumerate(row):
            text = norm_cell(value)
            if is_marker(text):
                marker_col = j
                marker_text = text
                break

        current = read_block(row, cols["current"], marker_col)
        recent = read_block(row, cols["recent"], marker_col)
        distant = read_block(row, cols["distant"], marker_col)

        if session_num is None and marker_text is None and current is None:
            continue  # blank / decorative row
        if session_num is None:
            anomalies.append(
                f"{path.name} / {sheet} row {idx}: row has content but no session number "
                f"(marker={marker_text!r}, current={current!r}) — skipped"
            )
            continue

        rows.append(
            RawRow(
                file=str(path.relative_to(CURRICULUM_DIR.parent)),
                sheet=sheet,
                row=idx,
                raw_session_number=session_num,
                marker_text=marker_text,
                marker_col=marker_col,
                current=current,
                recent=recent,
                distant=distant,
            )
        )
    return rows


# --------------------------------------------------------------------------
# Authoritative-tab selection
# --------------------------------------------------------------------------
# A workbook carries several tabs. Exactly ONE of them is the corrected,
# authoritative table; the others are abandoned drafts. The authoritative tab is
# the one that has been FORMATTED: its cells are colour-filled (200-900 filled
# cells), while a draft carries at most a handful of stray fills.
#
# The tab CANNOT be picked by name or by position: it is 'Sheet2' for levels
# 3-10 and 'Sheet1' for levels 1-2. Nor by xlsx tab colour: no workbook in the
# corpus sets sheetPr/tabColor. So we count fills, and demand a decisive winner.
MIN_AUTHORITATIVE_FILLS = 50   # every real tab in the corpus has >= 100
FILL_DOMINANCE = 5             # ... and beats every draft by far more than 5x


def fill_count(worksheet) -> int:
    """Cells carrying a real (non-white, non-transparent) background fill."""
    n = 0
    for row in worksheet.iter_rows():
        for cell in row:
            fill = cell.fill
            if fill is None or fill.patternType is None:
                continue
            colour = fill.fgColor
            if colour is None:
                continue
            if colour.type in ("theme", "indexed"):
                n += 1
            elif colour.type == "rgb" and colour.rgb not in (None, "00000000", "FFFFFFFF"):
                n += 1
    return n


def authoritative_sheet(path: Path) -> tuple[str, pd.DataFrame]:
    """The single colour-filled tab of a workbook, chosen by fill count."""
    workbook = load_workbook(path)
    try:
        counts = [(worksheet.title, fill_count(worksheet)) for worksheet in workbook.worksheets]
    finally:
        workbook.close()

    if not counts:
        raise Abort(f"{path.name}: workbook has no worksheets")

    counts.sort(key=lambda tc: tc[1], reverse=True)
    (winner, best), runners_up = counts[0], counts[1:]
    second = runners_up[0][1] if runners_up else 0

    if best < MIN_AUTHORITATIVE_FILLS:
        raise Abort(
            f"{path.name}: no colour-filled tab (best is {winner!r} with {best} filled "
            f"cells, below the {MIN_AUTHORITATIVE_FILLS} floor). Refusing to guess which "
            f"tab is authoritative. Fills: {counts}"
        )
    if second * FILL_DOMINANCE >= best:
        raise Abort(
            f"{path.name}: two tabs look equally formatted ({counts[0]} vs {counts[1]}); "
            f"cannot tell which is authoritative. Refusing to guess."
        )

    df = pd.read_excel(path, sheet_name=winner, header=None)
    if df.empty:
        raise Abort(f"{path.name} / {winner}: the colour-filled tab is empty")
    return winner, df


def hizb_from_filename(name: str) -> Optional[int]:
    match = re.search(r"الحزب\s*(?:ال|الــ|الـــ|الـ)?\s*(\d+)", name)
    if match:
        return int(match.group(1))
    match = re.search(r"الــ\s*(\d+)", name)
    if match:
        return int(match.group(1))
    return None


# --------------------------------------------------------------------------
# Juz assembly
# --------------------------------------------------------------------------
@dataclass
class JuzSource:
    level: int
    juz: int
    rows: list[RawRow]
    files: list[str]
    assembly: str = "single"  # single | merged | concatenated
    notes: list[str] = field(default_factory=list)
    half_hizbs: Optional[list[int]] = None  # teaching-ordered hizb numbers, concatenated only


def rows_signature(rows: list[RawRow]) -> list[tuple]:
    """What a row SAYS, independent of what it is numbered.

    The session number is deliberately left out: a redundant copy of a run of
    sessions is still a redundant copy when it is renumbered (L3 juz 24 is
    exactly that — see `_contained_in`).
    """
    return [
        (r.marker_text, json.dumps(r.current, ensure_ascii=False),
         json.dumps(r.recent, ensure_ascii=False), json.dumps(r.distant, ensure_ascii=False))
        for r in rows
    ]


def _contained_in(inner: list[RawRow], outer: list[RawRow]) -> bool:
    """True when every row of `inner` also appears, in order and adjacent, in `outer`."""
    a, b = rows_signature(inner), rows_signature(outer)
    if not a or len(a) >= len(b):
        return False
    return any(b[i:i + len(a)] == a for i in range(len(b) - len(a) + 1))


def _load_concatenated(
    level: int, juz: int, files: list[Path], parsed: dict[Path, list[RawRow]], notes: list[str]
) -> JuzSource:
    """Two workbooks that each renumber FROM SCRATCH: two hizbs, taught in turn.

    Their session numbers overlap but carry DIFFERENT content, so the two runs
    are two separate halves of the juz and are CONCATENATED in teaching order.
    A set-union keyed on session number would silently destroy one of them.
    """
    hizbs = {f: hizb_from_filename(f.name) for f in files}
    if any(h is None for h in hizbs.values()):
        raise Abort(
            f"Level {level} juz {juz}: two workbooks with overlapping-but-differing session "
            f"numbering, but no hizb number in the filenames to order them by: "
            f"{[f.name for f in files]}"
        )
    # Teaching order: level 1 ascends hizb, levels 2+ descend.
    ordered = sorted(files, key=lambda f: hizbs[f], reverse=(level != 1))
    # Corroboration, structural rather than textual (levels 3+ never name a juz
    # in their labels): a half carries the unit pair over its OWN hizb, and the
    # half taught LAST additionally carries the juz-tier pair (and, where the juz
    # is not the level's first, the cumulative pair). So exactly one half holds
    # more than one assessment, and it must be the one hizb order puts last.
    sards = {f: sum(1 for r in parsed[f] if kind_of(r) == "sard") for f in files}
    carriers = [f for f in files if sards[f] > 1]
    if len(carriers) != 1:
        raise Abort(
            f"Level {level} juz {juz}: expected exactly one half to carry more than its own "
            f"unit pair, found {len(carriers)}. sard counts: "
            f"{ {f.name: n for f, n in sards.items()} }"
        )
    if carriers[0] != ordered[-1]:
        raise Abort(
            f"Level {level} juz {juz}: teaching order by hizb says {ordered[-1].name!r} is "
            f"last, but the juz-tier pair lives in {carriers[0].name!r}. Refusing to guess."
        )
    notes.append(
        "CONCATENATED: two half-workbooks, each renumbering from scratch over different "
        f"content; teaching order {[hizbs[f] for f in ordered]} "
        f"(hizb {'ascending' if level == 1 else 'descending'}), corroborated by the "
        f"juz-tier pair living in the last half ({carriers[0].name!r})"
    )
    return JuzSource(
        level, juz,
        [r for f in ordered for r in parsed[f]],
        [f.name for f in ordered],
        "concatenated",
        notes,
        half_hizbs=[hizbs[f] for f in ordered],
    )


def _load_merged(
    level: int, juz: int, files: list[Path], parsed: dict[Path, list[RawRow]], notes: list[str]
) -> JuzSource:
    """Two workbooks holding DISJOINT halves of ONE continuous session numbering.

    e.g. 16-31 in one file and 1-15 in the other. The session number, not the
    file, is the order — the file with the LOWER hizb in its name routinely
    holds the HIGHER range.
    """
    rows = [r for f in files for r in parsed[f]]
    rows.sort(key=lambda r: r.raw_session_number)
    numbers = [r.raw_session_number for r in rows]
    if numbers != list(range(numbers[0], numbers[0] + len(numbers))):
        raise Abort(
            f"Level {level} juz {juz}: the two workbooks' session numbers are disjoint but do "
            f"not form one contiguous run: {numbers}"
        )
    ordered = sorted(files, key=lambda f: min(r.raw_session_number for r in parsed[f]))
    ranges = {
        f.name: (min(r.raw_session_number for r in parsed[f]),
                 max(r.raw_session_number for r in parsed[f]))
        for f in ordered
    }
    notes.append(
        "MERGED: two workbooks holding disjoint halves of one continuous numbering "
        f"{ranges}; ordered by session number"
    )
    return JuzSource(level, juz, rows, [f.name for f in ordered], "merged", notes)


def load_juz(level: int, juz: int, juz_dir: Path, anomalies: list[str]) -> JuzSource:
    """Assemble the sessions of one juz from the workbooks of its folder.

    Four layouts exist in the corpus, and we DETECT which one applies rather
    than keying off the level number:

      * single       — one workbook (levels 1, 8, 9, 10).
      * contained    — two workbooks, one of whose colour-filled tabs holds the
                       WHOLE juz and the other a renumbered copy of a run of it
                       (L3 juz 24 only). The containing tab is the juz; the copy
                       is redundant and is reported and dropped.
      * merged       — two workbooks whose session numbers are DISJOINT halves of
                       one continuous run (levels 4, 5, 6, 7).
      * concatenated — two workbooks whose session numbers OVERLAP but whose
                       content differs: two hizbs, each renumbering from scratch
                       (levels 2, 3).

    Anything else aborts. In particular we never take a set-union on session
    number: that would dedupe the distinct-but-identically-numbered sessions of
    the concatenated layout and destroy half the juz.
    """
    files = sorted(f for f in juz_dir.glob("*.xlsx") if not f.name.startswith("~$"))
    if not files:
        raise Abort(f"Level {level} juz {juz}: no .xlsx files in {juz_dir}")

    parsed: dict[Path, list[RawRow]] = {}
    for f in files:
        sheet, df = authoritative_sheet(f)
        parsed[f] = parse_sheet(f, sheet, df, anomalies)
        if not parsed[f]:
            raise Abort(f"Level {level} juz {juz}: {f.name} / {sheet} produced no rows")

    notes: list[str] = []

    if len(files) == 1:
        f = files[0]
        notes.append(
            f"SINGLE workbook; authoritative tab {sorted({r.sheet for r in parsed[f]})[0]!r}"
        )
        return JuzSource(level, juz, parsed[f], [f.name], "single", notes)

    if len(files) != 2:
        raise Abort(
            f"Level {level} juz {juz}: {len(files)} workbooks found, cannot decide layout: "
            f"{[f.name for f in files]}"
        )

    a, b = files
    numbers_a = [r.raw_session_number for r in parsed[a]]
    numbers_b = [r.raw_session_number for r in parsed[b]]
    for f, numbers in ((a, numbers_a), (b, numbers_b)):
        if len(set(numbers)) != len(numbers):
            raise Abort(
                f"Level {level} juz {juz}: {f.name} repeats a session number: {numbers}"
            )

    # One tab holding the whole juz, the other a renumbered copy of part of it.
    for whole, part in ((a, b), (b, a)):
        if _contained_in(parsed[part], parsed[whole]):
            anomalies.append(
                f"L{level} J{juz}: {part.name!r}'s colour-filled tab is a renumbered COPY of "
                f"{len(parsed[part])} of the {len(parsed[whole])} rows of {whole.name!r}'s, "
                f"which holds the whole juz on its own. The copy is redundant and is dropped; "
                f"concatenating it would have taught its sessions twice."
            )
            notes.append(
                f"CONTAINED: {whole.name!r} holds the whole juz; {part.name!r} is a renumbered "
                f"copy of {len(parsed[part])} of its rows and is dropped"
            )
            return JuzSource(level, juz, parsed[whole], [whole.name], "contained", notes)

    if not (set(numbers_a) & set(numbers_b)):
        return _load_merged(level, juz, files, parsed, notes)

    if rows_signature(parsed[a]) == rows_signature(parsed[b]):
        raise Abort(
            f"Level {level} juz {juz}: the two workbooks' authoritative tabs are IDENTICAL "
            f"({a.name!r} == {b.name!r}). That is neither a merge nor two hizbs; refusing to "
            f"guess which half of the juz is missing."
        )
    return _load_concatenated(level, juz, files, parsed, notes)


# --------------------------------------------------------------------------
# Classification & tiers
# --------------------------------------------------------------------------
def kind_of(row: RawRow) -> str:
    if row.marker_text:
        if row.marker_text.startswith("سرد"):
            return "sard"
        return "exam"
    if row.current or row.recent or row.distant:
        # A lesson with only review content is a real, taught consolidation session
        # (no new memorisation that day). It is reported, never dropped.
        return "lesson"
    return "anomaly"


def assessed_by(text: str) -> Optional[str]:
    if "المحفظ المتابع" in text:
        return "teacher"
    if "إدارة الحلقات" in text:
        return "supervisor"
    return None


def hizb_in_text(text: str) -> Optional[int]:
    match = HIZB_IN_TEXT_RE.search(text)
    return int(match.group(1)) if match else None


def juz_numbers_in_text(text: str) -> list[int]:
    if not JUZ_WORD_RE.search(text):
        return []
    tail = text[JUZ_WORD_RE.search(text).start():]
    tail = tail.split("على المحفظ")[0].split("من قِبل")[0]
    return [int(n) for n in JUZ_IN_TEXT_RE.findall(tail) if 1 <= int(n) <= 30]


@dataclass
class Pair:
    sard: int  # index into the juz's row list
    exam: int
    tier: str = ""
    unit_index: Optional[int] = None


def build_juz_sessions(
    src: JuzSource,
    juz_index_in_level: int,          # 0-based position in the level's teaching order
    juz_taught_so_far: list[int],     # including this juz
    errors: list[str],
    warnings: list[str],
    anomalies: list[str],
) -> list[dict]:
    level, juz = src.level, src.juz
    rows = src.rows
    kinds = [kind_of(r) for r in rows]

    for i, (r, k) in enumerate(zip(rows, kinds)):
        if k == "anomaly":
            anomalies.append(
                f"L{level} J{juz} {r.file} / {r.sheet} row {r.row}: session "
                f"{r.raw_session_number} has neither an assessment marker nor Quran content"
            )
            errors.append(f"L{level} J{juz}: un-classifiable row at {r.sheet}:{r.row}")
        elif k == "lesson" and not r.current:
            anomalies.append(
                f"L{level} J{juz} {r.file} / {r.sheet} row {r.row}: review-only lesson "
                f"(raw session {r.raw_session_number}) — no new مقرر, review content only"
            )

    # --- pairs: every assessment must be a (sard, exam) in that order --------
    pairs: list[Pair] = []
    i = 0
    while i < len(rows):
        if kinds[i] == "sard":
            if i + 1 < len(rows) and kinds[i + 1] == "exam":
                pairs.append(Pair(sard=i, exam=i + 1))
                i += 2
                continue
            errors.append(
                f"L{level} J{juz}: sard at {rows[i].sheet}:{rows[i].row} is not followed by an exam"
            )
            i += 1
            continue
        if kinds[i] == "exam":
            errors.append(
                f"L{level} J{juz}: exam at {rows[i].sheet}:{rows[i].row} is not preceded by a sard"
            )
        i += 1

    # --- structural tiers ---------------------------------------------------
    # A pair preceded by lesson rows is a `unit` pair. A pair immediately
    # following another pair is the `juz` pair; a further one is `cumulative`.
    exc = EXPECTED_EXCEPTIONS.get((level, juz))
    unit_counter = 0
    for p_i, pair in enumerate(pairs):
        previous_pair = pairs[p_i - 1] if p_i else None
        follows_pair = previous_pair is not None and previous_pair.exam == pair.sard - 1
        if not follows_pair:
            unit_counter += 1
            pair.tier = "unit"
            pair.unit_index = unit_counter
        elif previous_pair.tier == "unit":
            pair.tier = "juz"
        else:
            pair.tier = "cumulative"

    if exc and exc.get("second_pair_tier") and len(pairs) >= 2:
        if pairs[1].tier != exc["second_pair_tier"]:
            pairs[1].tier = exc["second_pair_tier"]
            warnings.append(
                f"L{level} J{juz}: applying documented exception — second pair re-tiered as "
                f"{exc['second_pair_tier']!r}. {exc['reason']}"
            )

    n_unit = sum(1 for p in pairs if p.tier == "unit")
    n_juz = sum(1 for p in pairs if p.tier == "juz")
    n_cum = sum(1 for p in pairs if p.tier == "cumulative")

    expect = exc or {"unit_pairs": 2, "juz_pairs": 1}
    if n_unit != expect.get("unit_pairs", 2) or n_juz != expect.get("juz_pairs", 1):
        msg = (
            f"L{level} J{juz}: found {n_unit} unit pair(s) and {n_juz} juz pair(s); "
            f"expected {expect.get('unit_pairs', 2)} / {expect.get('juz_pairs', 1)}"
        )
        (warnings if exc else errors).append(msg)

    # cumulative rules
    expected_cum = 0 if juz_index_in_level == 0 else 1
    if exc and "cumulative_pairs" in exc:
        expected_cum = exc["cumulative_pairs"]
    if n_cum != expected_cum:
        errors.append(
            f"L{level} J{juz}: {n_cum} cumulative pair(s), expected {expected_cum} "
            f"(juz #{juz_index_in_level + 1} taught in the level)"
        )

    # --- unit membership for lessons ---------------------------------------
    unit_pair_positions = [p.sard for p in pairs if p.tier == "unit"]

    def unit_of_row(i: int) -> Optional[int]:
        for n, pos in enumerate(unit_pair_positions, start=1):
            if i <= pos:
                return n
        return None

    # --- hizb labels --------------------------------------------------------
    # A unit is attributed to a hizb only where the SOURCE says so, i.e. where
    # its sard label names one ("سرد الحزب رقم 59 ..."). That is levels 1 and 2.
    # Level 2 additionally splits the juz across two per-hizb workbooks whose
    # labels are KNOWN-BAD (swapped between the halves), so there the structural
    # filename order is truth and the text is reported as a contradiction.
    # Levels 3-10 label their units by SURAH and never name a hizb -> null, even
    # where their workbooks happen to be named after one (the file names are
    # approximate; the taught unit is a surah, not a hizb).
    unit_pairs = [p for p in pairs if p.tier == "unit"]
    source_labels_units_by_hizb = any(
        hizb_in_text(rows[p.sard].marker_text or "") for p in unit_pairs
    )
    unit_hizb: dict[int, Optional[int]] = {1: None, 2: None}
    if src.half_hizbs and source_labels_units_by_hizb:
        for n, h in enumerate(src.half_hizbs, start=1):
            unit_hizb[n] = h
    else:
        for p in unit_pairs:
            unit_hizb[p.unit_index] = hizb_in_text(rows[p.sard].marker_text or "")

    for p in pairs:
        if p.tier != "unit":
            continue
        text_hizb = hizb_in_text(rows[p.sard].marker_text or "")
        structural = unit_hizb.get(p.unit_index)
        if text_hizb and structural and text_hizb != structural:
            warnings.append(
                f"L{level} J{juz}: HIZB LABEL CONTRADICTION — unit {p.unit_index} is structurally "
                f"hizb {structural} (file {rows[p.sard].file}) but its text says hizb {text_hizb}: "
                f"{rows[p.sard].marker_text!r}"
            )
            if level not in KNOWN_HIZB_LABEL_DEFECT_LEVELS:
                errors.append(
                    f"L{level} J{juz}: undocumented hizb label contradiction on unit {p.unit_index}"
                )

    # --- text cross-checks of the structural tier ---------------------------
    for p in pairs:
        text = rows[p.sard].marker_text or ""
        if p.tier == "unit":
            if JUZ_WORD_RE.search(text) or LEVEL_WORD_RE.search(text):
                errors.append(
                    f"L{level} J{juz}: unit-tier pair whose text is juz/level scoped: {text!r}"
                )
        elif p.tier == "juz":
            if hizb_in_text(text):
                errors.append(
                    f"L{level} J{juz}: juz-tier pair whose text names a hizb: {text!r}"
                )
            nums = juz_numbers_in_text(text)
            if nums and nums != [juz]:
                errors.append(
                    f"L{level} J{juz}: juz-tier pair whose text names juz {nums}: {text!r}"
                )
        elif p.tier == "cumulative":
            nums = juz_numbers_in_text(text)
            if nums and sorted(nums) != sorted(juz_taught_so_far):
                errors.append(
                    f"L{level} J{juz}: cumulative pair names juz {sorted(nums)} but the juz taught "
                    f"so far are {sorted(juz_taught_so_far)}: {text!r}"
                )

    pair_of_row: dict[int, Pair] = {}
    for p in pairs:
        pair_of_row[p.sard] = p
        pair_of_row[p.exam] = p

    # --- raw numbering sanity ----------------------------------------------
    # `single` and `merged` juz are ONE run and must increment by 1 throughout.
    # A `concatenated` juz is two runs, each restarting from scratch; each run is
    # checked on its own and the restart between them is expected, not a break.
    runs: list[list[int]] = []
    if src.assembly == "concatenated":
        for file_name in src.files:
            runs.append([r.raw_session_number for r in rows if r.file.endswith(file_name)])
    else:
        runs.append([r.raw_session_number for r in rows])

    for raw in runs:
        breaks = [
            (raw[i - 1], raw[i]) for i in range(1, len(raw)) if raw[i] != raw[i - 1] + 1
        ]
        if breaks:
            gap_note = EXPECTED_NUMBERING_GAPS.get((level, juz))
            msg = f"L{level} J{juz}: raw session numbers do not increment by 1: {breaks}"
            if gap_note:
                warnings.append(f"{msg} — documented: {gap_note}")
            else:
                errors.append(msg)

    # --- emit ---------------------------------------------------------------
    sessions = []
    for i, r in enumerate(rows):
        kind = kinds[i]
        pair = pair_of_row.get(i)
        unit_index = pair.unit_index if pair else unit_of_row(i)
        scope = None
        if pair:
            label = rows[pair.sard].marker_text if kind == "sard" else r.marker_text
            if pair.tier == "unit":
                juz_numbers = [juz]
            elif pair.tier == "juz":
                juz_numbers = [juz]
            else:
                juz_numbers = sorted(juz_taught_so_far)
            scope = {
                "tier": pair.tier,
                "label_ar": label,
                "hizb_number": hizb_in_text(label or ""),
                "juz_numbers": juz_numbers,
            }
        hizb = unit_hizb.get(unit_index) if unit_index else None

        sessions.append({
            "id": f"L{level}_J{juz}_S{i + 1}",
            "level_id": level,
            "juz_number": juz,
            "session_number": i + 1,
            "order_in_level": None,  # filled by the caller
            "kind": kind,
            "assessed_by": assessed_by(r.marker_text) if r.marker_text else None,
            "unit_index": unit_index if (kind == "lesson" or (pair and pair.tier == "unit")) else None,
            "hizb_number": hizb if (kind == "lesson" or (pair and pair.tier == "unit")) else None,
            "scope": scope,
            "current_level_content": r.current,
            "recent_review_content": r.recent,
            "distant_review_content": r.distant,
            "source": {
                "file": r.file,
                "sheet": r.sheet,
                "row": r.row,
                "raw_session_number": r.raw_session_number,
            },
        })

    # unit labels for levels.json
    src.notes.append(
        "tiers: " + ", ".join(f"{p.tier}{p.unit_index or ''}" for p in pairs)
    )
    return sessions


# --------------------------------------------------------------------------
# Derived sessions: تلقين
# --------------------------------------------------------------------------
# A تلقين session is one where the teacher recites the new passage TO the
# student and repeats it with him; the student memorizes nothing and is graded
# on nothing. It opens every unit — equivalently, it opens every level and
# follows every exam that is followed by new content.
#
# Nothing in the source spreadsheets is a تلقين row. These sessions are DERIVED,
# and say so: their `source` names the lesson they were derived from, never a
# file/sheet/row they did not come from.
def talqeen_of(lesson: dict) -> dict:
    """The تلقين session that introduces `lesson`.

    It teaches exactly the passage the student will memorize in `lesson`, and
    carries no review content, no scope and no assessor.

    `derived_from` cannot be filled in here: at this point `lesson["id"]` is
    still its PRE-renumbering id, and inserting this very talqeen shifts the
    id/session_number of every session from here on. `insert_talqeen_sessions`
    fills `derived_from` in once renumbering is done and every id is final.
    """
    return {
        "id": None,             # assigned by insert_talqeen_sessions
        "level_id": lesson["level_id"],
        "juz_number": lesson["juz_number"],
        "session_number": None,  # assigned by insert_talqeen_sessions
        "order_in_level": None,  # filled by the caller, as for every session
        "kind": "talqeen",
        "assessed_by": None,
        "unit_index": lesson["unit_index"],
        "hizb_number": lesson["hizb_number"],
        "scope": None,
        "current_level_content": lesson["current_level_content"],
        "recent_review_content": None,
        "distant_review_content": None,
        "source": {"derived_from": None},  # filled in below, once ids are final
    }


def insert_talqeen_sessions(sessions: list[dict]) -> list[dict]:
    """Open every unit of one juz with a تلقين, and renumber the juz 1..N.

    Renumbering is not bookkeeping: the document id and `session_number` are
    the session's identity, and `order_in_level` (assigned by the caller from
    the list this returns) is the sole advancement key.

    A talqeen is inserted immediately before the lesson it introduces, so once
    renumbering is done that lesson is always the very next session. Setting
    `derived_from` is therefore a second pass, run only after every id in the
    juz is final -- never the lesson's stale pre-renumbering id.
    """
    out: list[dict] = []
    opened: set[int] = set()
    for s in sessions:
        unit = s["unit_index"]
        if s["kind"] == "lesson" and unit is not None and unit not in opened:
            opened.add(unit)
            out.append(talqeen_of(s))
        out.append(s)

    level = out[0]["level_id"]
    juz = out[0]["juz_number"]
    for i, s in enumerate(out):
        s["session_number"] = i + 1
        s["id"] = f"L{level}_J{juz}_S{i + 1}"

    for i, s in enumerate(out):
        if s["kind"] == "talqeen":
            s["source"]["derived_from"] = out[i + 1]["id"]

    return out


# --------------------------------------------------------------------------
# Driver
# --------------------------------------------------------------------------
def extract() -> tuple[dict, dict, list[str], list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []
    anomalies: list[str] = []

    levels: dict[int, dict] = {}
    sessions_by_level: dict[int, list[dict]] = {}

    for folder, level in sorted(LEVEL_FOLDERS.items(), key=lambda kv: kv[1]):
        level_dir = CURRICULUM_DIR / folder
        if not level_dir.is_dir():
            raise Abort(f"Level {level}: folder not found: {level_dir}")

        order = juz_teaching_order(level)
        juz_dirs: dict[int, Path] = {}
        for d in level_dir.iterdir():
            if not d.is_dir():
                continue
            match = re.search(r"الجزء\s*ال[ـ\s]*(\d+)", d.name)
            if not match:
                raise Abort(f"Level {level}: cannot read a juz number from folder {d.name!r}")
            juz_dirs[int(match.group(1))] = d

        if sorted(juz_dirs) != sorted(order):
            raise Abort(
                f"Level {level}: expected juz {sorted(order)}, found {sorted(juz_dirs)}"
            )

        level_sessions: list[dict] = []
        juz_meta: list[dict] = []
        taught: list[int] = []

        for pos, juz in enumerate(order):
            src = load_juz(level, juz, juz_dirs[juz], anomalies)
            taught.append(juz)
            juz_sessions = build_juz_sessions(
                src, pos, list(taught), errors, warnings, anomalies
            )
            # A تلقين opens every unit. Inserted BEFORE order_in_level is
            # assigned, so the level ordering, the level session_count and
            # first_order_in_level all account for them without a second pass.
            juz_sessions = insert_talqeen_sessions(juz_sessions)
            first_order = len(level_sessions) + 1
            for s in juz_sessions:
                s["order_in_level"] = len(level_sessions) + 1
                level_sessions.append(s)

            unit_labels = [
                s["scope"]["label_ar"]
                for s in juz_sessions
                if s["scope"] and s["scope"]["tier"] == "unit" and s["kind"] == "sard"
            ]
            hizbs = sorted({s["hizb_number"] for s in juz_sessions if s["hizb_number"]})
            juz_meta.append({
                "juz_number": juz,
                "session_count": len(juz_sessions),
                "unit_labels": unit_labels,
                "hizb_numbers": hizbs or None,
                "first_order_in_level": first_order,
                "source_files": src.files,
                "notes": src.notes,
            })

        levels[level] = {
            "id": level,
            "name_ar": LEVEL_NAMES_AR[level],
            "name_en": LEVEL_NAMES_EN[level],
            "order": level,
            "juz_numbers": order,
            "session_count": len(level_sessions),
            "juz": juz_meta,
        }
        sessions_by_level[level] = level_sessions

    # ---- corpus-wide validation -------------------------------------------
    CONTENT_BLOCKS = (
        "current_level_content", "recent_review_content", "distant_review_content",
    )
    seen_ids: set[str] = set()
    for level, sessions in sessions_by_level.items():
        for s in sessions:
            for block in CONTENT_BLOCKS:
                passage = s[block]
                if not passage:
                    continue
                # A passage that starts and ends in the same surah cannot run
                # BACKWARDS. Every such inversion in the old output came from
                # reading a draft tab; the authoritative tabs have none, so this
                # is an error, not a tolerated anomaly.
                if (
                    normalise_surah(passage["from_surah"])
                    == normalise_surah(passage["to_surah"])
                    and passage["to_verse"] < passage["from_verse"]
                ):
                    errors.append(
                        f"{s['id']} {block}: inverted range — {passage['from_surah']} "
                        f"{passage['from_verse']} : {passage['to_verse']}"
                    )
                # A name that is not one of the 114 surahs is a TYPO IN THE
                # SOURCE. Reported, stored verbatim, never corrected here.
                for key in ("from_surah", "to_surah"):
                    if not is_known_surah(passage[key]):
                        anomalies.append(
                            f"{s['id']} {block}.{key}: {passage[key]!r} is not the name of any "
                            f"surah — SOURCE TYPO, stored verbatim (from {s['source'].get('file')} "
                            f"/ {s['source'].get('sheet')} row {s['source'].get('row')})"
                        )
        if len(levels[level]["juz"]) != 3:
            errors.append(f"L{level}: has {len(levels[level]['juz'])} juz, expected 3")
        for juz_meta in levels[level]["juz"]:
            if juz_meta["session_count"] < 1:
                errors.append(f"L{level} J{juz_meta['juz_number']}: no sessions")
        by_juz: dict[int, list[int]] = {}
        for s in sessions:
            if s["id"] in seen_ids:
                errors.append(f"duplicate document id {s['id']}")
            seen_ids.add(s["id"])
            by_juz.setdefault(s["juz_number"], []).append(s["session_number"])
            if s["kind"] == "lesson" and not any((
                s["current_level_content"], s["recent_review_content"], s["distant_review_content"],
            )):
                errors.append(f"{s['id']}: lesson with no Quran content at all")
            if s["kind"] in ("sard", "exam") and not s["scope"]:
                errors.append(f"{s['id']}: assessment without a scope")
            if s["kind"] == "talqeen":
                if not s["source"].get("derived_from"):
                    errors.append(f"{s['id']}: talqeen without a derived_from")
                if not s["current_level_content"]:
                    errors.append(f"{s['id']}: talqeen teaches no passage")
                if s["scope"] or s["assessed_by"]:
                    errors.append(f"{s['id']}: talqeen is not an assessment")
            elif not s["source"] or s["source"]["row"] is None:
                errors.append(f"{s['id']}: missing source provenance")
        for juz, nums in by_juz.items():
            if nums != list(range(1, len(nums) + 1)):
                errors.append(f"L{level} J{juz}: session numbers are not dense 1..N: {nums}")

        # Every unit opens with a تلقين that teaches the passage of the lesson
        # it precedes. A unit that opens with anything else is a derivation bug.
        by_order = sorted(sessions, key=lambda s: s["order_in_level"])
        units_opened: set[tuple[int, int]] = set()
        for i, s in enumerate(by_order):
            unit = s["unit_index"]
            if unit is None:
                continue
            key = (s["juz_number"], unit)
            if key in units_opened:
                continue
            units_opened.add(key)
            if s["kind"] != "talqeen":
                errors.append(
                    f"L{level} J{s['juz_number']} unit {unit}: opens with "
                    f"{s['kind']} ({s['id']}), not a talqeen"
                )
                continue
            following = by_order[i + 1]
            if s["current_level_content"] != following["current_level_content"]:
                errors.append(
                    f"{s['id']}: talqeen does not teach the passage of "
                    f"{following['id']}"
                )
            if s["source"].get("derived_from") != following["id"]:
                errors.append(
                    f"{s['id']}: derived_from is {s['source'].get('derived_from')!r}, "
                    f"expected the id of the session it introduces ({following['id']!r})"
                )

    return levels, sessions_by_level, errors, warnings, anomalies


def print_report(levels, sessions_by_level, errors, warnings, anomalies) -> None:
    print("\n" + "=" * 96)
    print(f"{'level':>5} {'juz':>4} {'sessions':>9} {'talqeen':>8} {'lessons':>8} {'sard':>5} {'exam':>5}  tiers")
    print("-" * 96)
    for level in sorted(levels):
        for juz_meta in levels[level]["juz"]:
            juz = juz_meta["juz_number"]
            ss = [s for s in sessions_by_level[level] if s["juz_number"] == juz]
            tiers = [
                f"{s['scope']['tier']}{s['unit_index'] or ''}"
                for s in ss if s["kind"] == "sard"
            ]
            print(
                f"{level:>5} {juz:>4} {len(ss):>9} "
                f"{sum(1 for s in ss if s['kind'] == 'talqeen'):>8} "
                f"{sum(1 for s in ss if s['kind'] == 'lesson'):>8} "
                f"{sum(1 for s in ss if s['kind'] == 'sard'):>5} "
                f"{sum(1 for s in ss if s['kind'] == 'exam'):>5}  {', '.join(tiers)}"
            )
        print(f"{'':>5} {'':>4} {levels[level]['session_count']:>9}  (level total)")
    print("=" * 96)

    print(f"\nanomalies ({len(anomalies)}):")
    for a in anomalies:
        print(f"  - {a}")
    print(f"\nwarnings ({len(warnings)}):")
    for w in warnings:
        print(f"  ! {w}")
    print(f"\nerrors ({len(errors)}):")
    for e in errors:
        print(f"  X {e}")
    print()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--write", action="store_true", help="emit JSON after a clean validation")
    args = parser.parse_args()

    try:
        levels, sessions_by_level, errors, warnings, anomalies = extract()
    except Abort as exc:
        print(f"ABORT: {exc}", file=sys.stderr)
        return 2

    print_report(levels, sessions_by_level, errors, warnings, anomalies)

    total = sum(len(v) for v in sessions_by_level.values())
    report = {
        "generated_from": str(CURRICULUM_DIR),
        "levels": len(levels),
        "total_sessions": total,
        "per_level": {
            str(level): {
                "session_count": levels[level]["session_count"],
                "juz": {
                    str(j["juz_number"]): {
                        "session_count": j["session_count"],
                        "source_files": j["source_files"],
                        "notes": j["notes"],
                        "hizb_numbers": j["hizb_numbers"],
                    }
                    for j in levels[level]["juz"]
                },
            }
            for level in sorted(levels)
        },
        "expected_exceptions": {
            f"L{lvl}_J{juz}": exc for (lvl, juz), exc in EXPECTED_EXCEPTIONS.items()
        },
        "expected_numbering_gaps": {
            f"L{lvl}_J{juz}": note for (lvl, juz), note in EXPECTED_NUMBERING_GAPS.items()
        },
        "anomalies": anomalies,
        "warnings": warnings,
        "errors": errors,
        "passed": not errors,
    }

    if errors:
        print(f"VALIDATION FAILED with {len(errors)} error(s). Nothing was written.", file=sys.stderr)
        if args.write:
            return 1
        return 1

    print(f"VALIDATION PASSED — {total} sessions across {len(levels)} levels.")
    if not args.write:
        print("(dry run; pass --write to emit JSON)")
        return 0

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUTPUT_DIR / "levels.json").write_text(
        json.dumps([levels[l] for l in sorted(levels)], ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    for level, sessions in sessions_by_level.items():
        (OUTPUT_DIR / f"sessions_level_{level}.json").write_text(
            json.dumps(sessions, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
    (OUTPUT_DIR / "metadata.json").write_text(
        json.dumps(
            {
                "source": "curriculum/ (read-only spreadsheets)",
                "extractor": "tools/curriculum/extract_curriculum.py",
                "levels": len(levels),
                "total_sessions": total,
                "sessions_per_level": {
                    str(l): levels[l]["session_count"] for l in sorted(levels)
                },
                "schema_version": 3,
            },
            ensure_ascii=False,
            indent=2,
        ) + "\n",
        encoding="utf-8",
    )
    (OUTPUT_DIR / "validation_report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(f"Wrote {OUTPUT_DIR}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
